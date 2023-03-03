# importa biblioteca
import fitz
import os
import re
from PIL import Image
import cv2
import pytesseract
import pandas as pd
from openpyxl import load_workbook
from os.path import exists
import os.path

file_path_excel = "Projeto_Revisao_Cartas_01032023.xlsx"
wb = load_workbook(file_path_excel)
sheet = wb.active

max_row = sheet.max_row
sair_pdf = False
valorRepetido = ""
anterior = ""
preenche_direita = False
for pdf in range(2, max_row+1):

    filtro = sheet.cell(row=pdf, column=6).value
    if not re.search(r"AUTO", str(filtro), re.IGNORECASE):
        continue

    if anterior == sheet.cell(row=pdf, column=3).value:
        if preenche_direita == False:
            sheet.cell(row=pdf, column=7).value = valorRepetido
        else:
            sheet.cell(row=pdf, column=9).value = valorRepetido
        continue

    valorRepetido = ""
    anterior = sheet.cell(row=pdf, column=3).value

    print(sheet.cell(row=pdf, column=3).value)
    sair_pdf = False
    # Define path to PDF file
    file_path = "pdf/" + sheet.cell(row=pdf, column=3).value + ".pdf"

    # Define path for saved images
    images_path = 'images/'

    if exists(file_path):
        # Open PDF file
        pdf_file = fitz.open(file_path)
    else:
        sheet.cell(row=pdf, column=9).value = "CARTA INEXISTENTE"
        valorRepetido = "CARTA INEXISTENTE"
        preenche_direita = True
        continue

    # Get the number of pages in PDF file
    page_nums = len(pdf_file)

    # Create empty list to store images information
    images_list = []

    # Extract all images information from each page
    for page_num in range(page_nums):
        page_content = pdf_file[page_num]
        images_list.extend(page_content.get_images())

    # Raise error if PDF has no images
    if len(images_list) == 0:
        continue

    # Save all the extracted images
    for i, img in enumerate(images_list, start=1):
        if sair_pdf == True:
            break
        # Extract the image object number
        xref = img[0]
        # Extract image
        base_image = pdf_file.extract_image(xref)
        # Store image bytes
        image_bytes = base_image['image']
        # Store image extension
        image_ext = base_image['ext']
        # Generate image file name
        image_name = str(i) + '.' + image_ext
        # Save image
        with open(os.path.join(images_path, image_name), 'wb') as image_file:
            image_file.write(image_bytes)
            image_file.close()

        pytesseract.pytesseract.tesseract_cmd = "C:\Program Files (x86)\Tesseract-OCR\Tesseract.exe"
       


        resultado = ""
        for path in os.listdir("images/"):
    
            img = cv2.imread("images/" + path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            blur = cv2.GaussianBlur(gray, (3,3), 0)
            thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
            opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
            invert = 255 - opening
            resultado = pytesseract.image_to_string(invert, lang="eng")

            if re.search(r"devo|eletr", resultado, re.IGNORECASE):
                sheet.cell(row=pdf, column=7).value = "TEM CEDO"
                os.remove("images/" + path)
                sair_pdf = True
                valorRepetido = "TEM CEDO"
                preenche_direita = False
                break
            else:
                sheet.cell(row=pdf, column=7).value = "NÃO TEM CEDO"
                valorRepetido = "NÃO TEM CEDO"
                preenche_direita = False
            os.remove("images/" + path)

wb.save(file_path_excel)
